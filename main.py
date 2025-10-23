from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

# def gender_select
def gender_select():
    global gender
    value = radio.get()
    if value == 1:
        gender = 'Мужской'
    else:
        gender = 'Женский'


# Exit
def Exit():
    root.destroy()


# Registration No:
def regisration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    row = sheet.max_row
    max_row_value = sheet.cell(row=row, column=1).value
    try:
        Registration.set(int(max_row_value) + 1)
    except:
        Registration.set('1')


def Clear():
    global image
    global img
    regisration_no()
    First_name.set('')
    Last_name.set('')
    age.set('')
    kyrs.set('Выберите курс')
    skill.set('')
    Father_name.set('')
    Father_job.set('')
    Mather_name.set('')
    Mather_job.set('')
    save_button.config(state='normal')
    image = PhotoImage(file="images/upload photo.png")
    img_lbl.config(image=image)
    img_lbl.image = image
    img = ''
    

def Save():
    R1 = Registration.get()
    D1 = Date.get()
    Fst_n1 = First_name.get()
    L_n1 = Last_name.get()
    A1 = age.get()
    try:
        G1 = gender
    except:
        messagebox.showerror('error', 'Не выбран пол!')
    K1 = kyrs.get()
    S1 = skill.get()
    F_n1 = Father_name.get()
    F_j1 = Father_job.get()
    M_n1 = Mather_name.get()
    M_j1 = Mather_job.get()
    if Fst_n1 == '' or L_n1 == '' or A1 == '' or len(K1) > 1 or S1 == '' or F_n1 == '' or F_j1 == '' or M_n1 == '' or M_j1 == '':
        messagebox.showerror('error', 'Не все данные введены!')
    else:
        file = openpyxl.load_workbook('Student_data.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=D1)
        sheet.cell(column=3, row=sheet.max_row, value=Fst_n1)
        sheet.cell(column=4, row=sheet.max_row, value=L_n1)
        sheet.cell(column=5, row=sheet.max_row, value=A1)
        sheet.cell(column=6, row=sheet.max_row, value=G1)
        sheet.cell(column=7, row=sheet.max_row, value=K1)
        sheet.cell(column=8, row=sheet.max_row, value=S1)
        sheet.cell(column=9, row=sheet.max_row, value=F_n1)
        sheet.cell(column=10, row=sheet.max_row, value=F_j1)
        sheet.cell(column=11, row=sheet.max_row, value=M_n1)
        sheet.cell(column=12, row=sheet.max_row, value=M_j1)
        file.save(r'Student_data.xlsx')
        try:
            img.save("students_images/" + str(R1) + ".jpg")
        except:
            messagebox.showinfo('info', 'Изображение пользователя не доступно!')
        messagebox.showinfo('info', 'Данные студента успешно сохранены!')
        Clear()
        regisration_no()


def showfile():
    global filename
    global img    
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Выберите файл изображения",
                                          filetypes=(("JPG File", "*.jpg"),
                                                    ("PNG File", "*.png"),
                                                    ("All files", "*.txt"),))
    img = (Image.open(filename))
    resized_image = img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    img_lbl.config(image=photo2)
    img_lbl.image = photo2
    

def search():
    text = Search.get()
    Clear()
    save_button.config(state='disabled')
    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active
    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            # reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
    try:
        # print(str(name))
        pass
    except:
        messagebox.showerror("Invalid", "Этот номер не найден!")
    x1 = sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value
    Registration.set(x1)
    Date.set(x2)
    First_name.set(x3)
    Last_name.set(x4)
    age.set(x5)
    if x6 == 'Мужской':
        R1.select()
    else:
        R2.select()
    kyrs.set(x7)
    skill.set(x8)
    Father_name.set(x9)
    Father_job.set(x10)
    Mather_name.set(x11)
    Mather_job.set(x12)
    img = (Image.open("students_images/" + str(x1) +".jpg"))
    resized_image = img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    img_lbl.config(image=photo2)
    img_lbl.image = photo2


def Update():
    R1 = Registration.get()
    D1 = Date.get()
    Fst_n1 = First_name.get()
    L_n1 = Last_name.get()
    A1 = age.get()
    gender_select()
    G1 = gender    
    K1 = kyrs.get()
    S1 = skill.get()
    F_n1 = Father_name.get()
    F_j1 = Father_job.get()
    M_n1 = Mather_name.get()
    M_j1 = Mather_job.get()
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    for row in sheet.rows:
        if row[0].value == R1:
            name = row[0]
            reg_number = str(name)[15:-1]
    sheet.cell(row=int(reg_number), column=1, value=R1)
    sheet.cell(row=int(reg_number), column=2, value=D1)
    sheet.cell(row=int(reg_number), column=3, value=Fst_n1)
    sheet.cell(row=int(reg_number), column=4, value=L_n1)
    sheet.cell(row=int(reg_number), column=5, value=A1)
    sheet.cell(row=int(reg_number), column=6, value=G1)
    sheet.cell(row=int(reg_number), column=7, value=K1)
    sheet.cell(row=int(reg_number), column=8, value=S1)
    sheet.cell(row=int(reg_number), column=9, value=F_n1)
    sheet.cell(row=int(reg_number), column=10, value=F_j1)
    sheet.cell(row=int(reg_number), column=11, value=M_n1)
    sheet.cell(row=int(reg_number), column=12, value=M_j1)
    file.save(r'Student_data.xlsx')
    try:
        img.save("students_images/" + str(R1) + ".jpg")
    except:
        pass
    messagebox.showinfo('info', 'Данные студента успешно обновлены!')
    Clear()


root = Tk()
root.title("Список студентов")
root.geometry("1250x700+230+100")
root.config(bg=background)

file = pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = '№'
    sheet['B1'] = 'Дата регистрации'
    sheet['C1'] = 'Имя'
    sheet['D1'] = 'Фамилия'
    sheet['E1'] = 'Возраст'
    sheet['F1'] = 'Пол'
    sheet['G1'] = 'Курс'
    sheet['H1'] = 'Факультет'
    sheet['I1'] = 'Имя отца'
    sheet['J1'] = 'Имя матери'
    sheet['K1'] = 'Место работы отца'
    sheet['L1'] = 'Место работы матери'
    file.save('Student_data.xlsx')

# Top Frame
Label(root, text="andr_vol@mail.ru", width=10, height=3, bg="#f0687c", anchor="e").pack(side=TOP, fill=X)
Label(root, text="Список студентов", width=10, height=2, bg="#c36464", fg='#fff', font='arial 20 bold').pack(side=TOP, fill=X)

# Search Box
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font="arial 20").place(x=820, y=70)
image_icon_srch = PhotoImage(file="images/search.png")
Srch = Button(root, text="Найти", compound=LEFT, image=image_icon_srch, width=123, bg="#68ddfa", font="arial 13 bold", command=search)
Srch.place(x=1060, y=66)

image_icon_load = PhotoImage(file="images/Layer 4.png")
Update_button = Button(root, image=image_icon_load, bg="#c36464", command=Update)
Update_button.place(x=110, y=64)

# Registration and Date
Label(root, text="Регистрационный №:", font="arial 13", fg=framebg, bg=background).place(x=30, y=150)
Label(root, text="Дата:", font="arial 13", fg=framebg, bg=background).place(x=500, y=150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 12")
reg_entry.place(x=215, y=150)

regisration_no() # функция автоматически подставляет регистрационный номер

today = date.today()
d1 = today.strftime("%d/%m/%Y")

date_entry = Entry(root, textvariable=Date, width=15, font="arial 12")
date_entry.place(x=555, y=150)
Date.set(d1)

# Student details

obj = LabelFrame(root, text="Данные студента", font=20, bd=2, width=900, bg=framebg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Имя: ", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Фамилия: ", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Дата рождения: ", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)

Label(obj, text="Пол: ", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Курс: ", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Факультет: ", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

First_name = StringVar()
First_name_entry = Entry(obj, textvariable=First_name, width=20, font="arial 12")
First_name_entry.place(x=170, y=50)

Last_name = StringVar()
Last_name_entry = Entry(obj, textvariable=Last_name, width=20, font="arial 12")
Last_name_entry.place(x=170, y=100)

age = StringVar()
age_entry = Entry(obj, textvariable=age, width=20, font="arial 12")
age_entry.place(x=170, y=150)

radio = IntVar()
R1 = Radiobutton(obj, text="Муж", variable=radio, value=1, bg=framebg, fg=framefg, command=gender_select)
R1.place(x=600, y=50)

R2 = Radiobutton(obj, text="Жен", variable=radio, value=2, bg=framebg, fg=framefg, command=gender_select)
R2.place(x=700, y=50)

kyrs = Combobox(obj, values=['1', '2', '3', '4', '5'], font="Roboto 12", width=15, state="r")
kyrs.place(x=600, y=100)
kyrs.set('Выберите курс')

skill = StringVar()
skill_entry = Entry(obj, textvariable=skill, width=20, font="arial 12")
skill_entry.place(x=600, y=150)

# Parents details

obj2 = LabelFrame(root, text="Данные родителей", font=20, bd=2, width=900, bg=framebg, height=220, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Отец: ", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj2, text="Место работы: ", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)

Label(obj2, text="Мать: ", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj2, text="Место работы: ", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)

Father_name = StringVar()
Father_name_entry = Entry(obj2, textvariable=Father_name, width=20, font="arial 12")
Father_name_entry.place(x=170, y=50)

Father_job = StringVar()
Father_job_entry = Entry(obj2, textvariable=Father_job, width=20, font="arial 12")
Father_job_entry.place(x=170, y=100)

Mather_name = StringVar()
Mather_name_entry = Entry(obj2, textvariable=Mather_name, width=20, font="arial 12")
Mather_name_entry.place(x=620, y=50)

Mather_job = StringVar()
Mather_job_entry = Entry(obj2, textvariable=Mather_job, width=20, font="arial 12")
Mather_job_entry.place(x=620, y=100)

# Image
I_frame = Frame(root, bd=3, bg='black', width=195, height=190, relief=GROOVE)
I_frame.place(x=1000, y=150)

image = PhotoImage(file="images/upload photo.png")
img_lbl = Label(I_frame, bg='black', image=image)
img_lbl.place(x=0, y=0)

# Button
Button(root, text='Загрузить', width=17, height=2, font='arial 14 bold', bg='lightblue', command=showfile).place(x=1000, y=370)

save_button = Button(root, text='Сохранить', width=17, height=2, font='arial 14 bold', bg='lightgreen', command=Save)
save_button.place(x=1000, y=450)

Button(root, text='Сбросить', width=17, height=2, font='arial 14 bold', bg='lightpink', command=Clear).place(x=1000, y=530)

Button(root, text='Выход', width=17, height=2, font='arial 14 bold', bg='gray', command=Exit).place(x=1000, y=610)

root.mainloop()
